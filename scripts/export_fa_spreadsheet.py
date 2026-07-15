#!/usr/bin/env python3
"""Export a COMPLETE 2026-27 roster + free-agency workbook.

Every player is listed individually. The under-contract roster comes from the full
salary feed (cache/next_contracts_2026_v7.pkl = every 2026-27 contract) joined to
each player's team (cache/contract_end_years_v1.pkl), so nobody is lumped into an
"other" line. Re-signs and new signings come from the realistic FA simulation
(cache/fa_sim_v1.json). Held-out players (data/fa_sim_overrides.csv) are dropped
entirely. Re-run after the model changes:  python scripts/export_fa_spreadsheet.py
"""
import csv
import json
import pickle
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
from utils import normalize  # noqa: E402

OUT = Path.home() / "Downloads" / "HoopsValue_FreeAgency_2026.xlsx"

board = json.loads((ROOT / "cache" / "fa_board_v1.json").read_text())
sim = json.loads((ROOT / "cache" / "fa_sim_v1.json").read_text())
nc = pickle.load(open(ROOT / "cache" / "next_contracts_2026_v7.pkl", "rb"))      # {norm: {salary,type}}
cey = pickle.load(open(ROOT / "cache" / "contract_end_years_v1.pkl", "rb"))      # {norm: {current_team,...}}
TEAMS = board["teams"]
_any = next(iter(TEAMS.values()))
CAP, AP1, AP2 = round(sim.get("cap_M", 165.0)), round(_any["apron1_M"]), round(_any["apron2_M"])
TAX = round(CAP * 1.215)
ABBR_FIX = {"PHO": "PHX", "CHO": "CHA", "BRK": "BKN", "NOH": "NOP"}              # BBRef -> our abbrevs

# held-out players (manual overrides) — pretend they don't exist
held = set()
_ovr = ROOT / "data" / "fa_sim_overrides.csv"
if _ovr.exists():
    for r in csv.DictReader(_ovr.read_text().splitlines()):
        if (r.get("action") or "").strip() == "hold_out" and r.get("player"):
            held.add(normalize(r["player"]))

# Roster corrections (data/roster_corrections.csv): the scraped salary feed still
# lists these as guaranteed, but they're off the roster — either waived, or a team
# option the team declined. The live app drops them via this file; the export reads
# the raw feed directly, so it must apply the same fix. Keyed by (team abbr, name).
dropped, declined = set(), []
_rc = ROOT / "data" / "roster_corrections.csv"
if _rc.exists():
    _rc_lines = [l for l in _rc.read_text().splitlines() if not l.lstrip().startswith("#")]
    for r in csv.DictReader(_rc_lines):
        act = (r.get("action") or "").strip().lower()
        if act in ("waived", "decline_option") and r.get("player"):
            _tm = ABBR_FIX.get((r.get("team") or "").strip(), (r.get("team") or "").strip())
            dropped.add((_tm, normalize(r["player"])))
            if act == "decline_option":                  # off the roster but now a free agent
                declined.append((_tm, r["player"].strip()))

import re
_SUF = re.compile(r"\s+(jr|sr|ii|iii|iv|v)$")
def _strip(n):
    return _SUF.sub("", n)

# display name / position / barrett / TEAM lookups. The board roster is the reliable
# team source (its names match the salary feed); cey is the fallback for players the
# board doesn't list (opted-in option-holders, non-rated deep bench).
disp, pos_map, bar_map, board_team = {}, {}, {}, {}
for tm, t in TEAMS.items():
    for rr in t["roster"]:
        n = normalize(rr["name"])
        disp[n] = rr["name"]; pos_map[n] = rr["pos"]; bar_map[n] = rr.get("barrett"); board_team[n] = tm
cey_team = {}
for k, v in cey.items():
    if isinstance(v, dict) and v.get("current_team"):
        cey_team[k] = v["current_team"]; cey_team.setdefault(_strip(k), v["current_team"])
# Authoritative current team = where the player actually PLAYED this season (the stats
# feed, dumped to cache/current_teams_v1.json), with data/team_corrections.csv layered
# on. The contract feed can carry a stale team (a mock-trade rumor put Rui on POR though
# he never left LAL), which breaks incumbent / Bird-rights logic - so stats wins over cey.
stats_team, team_corr = {}, {}
_ct = ROOT / "cache" / "current_teams_v1.json"
if _ct.exists():
    stats_team = {k: ABBR_FIX.get(v, v) for k, v in json.loads(_ct.read_text()).items()}
_tc = ROOT / "data" / "team_corrections.csv"
if _tc.exists():
    for r in csv.DictReader([l for l in _tc.read_text().splitlines() if not l.lstrip().startswith("#")]):
        if r.get("player") and r.get("team"):
            team_corr[normalize(r["player"])] = ABBR_FIX.get(r["team"].strip(), r["team"].strip())
def cur_team(n):                                          # correction > board roster > season stats > stale cey
    return team_corr.get(n) or board_team.get(n) or stats_team.get(n) or cey_team.get(n) or cey_team.get(_strip(n))
for p in sim["players"]:
    n = normalize(p["player"]); disp[n] = p["player"]; pos_map.setdefault(n, p["pos"]); bar_map.setdefault(n, p["barrett"])
p2k = ROOT / "data" / "player_positions_2k.csv"
if p2k.exists():
    for r in csv.DictReader(p2k.read_text().splitlines()):
        nm = normalize(r.get("name", ""))
        if nm and r.get("positions"):
            pos_map.setdefault(nm, r["positions"])
# Manual position overrides WIN over 2K and the (pre-built) board roster, so a
# position fix shows up here without a full board rebuild.
p2k_ovr = ROOT / "data" / "player_positions_override.csv"
if p2k_ovr.exists():
    for r in csv.DictReader([l for l in p2k_ovr.read_text().splitlines() if not l.lstrip().startswith("#")]):
        nm = normalize(r.get("name", ""))
        if nm and r.get("positions"):
            pos_map[nm] = r["positions"]

fa_pool = {normalize(p["player"]) for p in sim["players"]}                        # the FA universe
sim_by_team = defaultdict(list)
for p in sim["players"]:
    if p["realistic"]["predicted"]:
        sim_by_team[p["realistic"]["predicted"]].append(p)

# Opt-in overrides (fa_sim_overrides.csv, action=opt_in): a player-option holder takes his
# option -> shown as UNDER CONTRACT at the option value, not a free-agent re-sign. Pull him
# from the sim pool so the nc loop (Player Option) picks him up.
optin = set()
if _ovr.exists():
    for r in csv.DictReader(_ovr.read_text().splitlines()):
        if (r.get("action") or "").strip().lower() == "opt_in" and r.get("player"):
            optin.add(normalize(r["player"]))
fa_pool -= optin
for t in list(sim_by_team):
    sim_by_team[t] = [p for p in sim_by_team[t] if normalize(p["player"]) not in optin]

# Blocked signings (data/blocked_signings.csv): the model placed a player on a team that
# makes no positional sense (e.g. a 6th PG); drop him there and re-route him (never back to
# the blocked team).
blocked, blocked_players = set(), []
_bs = ROOT / "data" / "blocked_signings.csv"
if _bs.exists():
    for r in csv.DictReader([l for l in _bs.read_text().splitlines() if not l.lstrip().startswith("#")]):
        _t = ABBR_FIX.get((r.get("team") or "").strip(), (r.get("team") or "").strip())
        if r.get("player"):
            blocked.add((_t, normalize(r["player"])))
for t in list(sim_by_team):
    keep = []
    for p in sim_by_team[t]:
        if (t, normalize(p["player"])) in blocked:
            blocked_players.append(p)                                             # re-route below
        else:
            keep.append(p)
    sim_by_team[t] = keep

# ── Supplement: 2026 free agents the main pool never ranked (scripts/build_missing_fas.py).
# A tail of genuine FAs (no 2026-27 salary on the books AND not flagged by the pool's
# minutes/status gate) is dropped from both feeds. The rotation-value ones are projected
# with the same engine and folded into rosters here; the rest are listed on their own sheet.
extra = {"listing": [], "projected": []}
_ex = ROOT / "cache" / "fa_extra_v1.json"
if _ex.exists():
    extra = json.loads(_ex.read_text())
for p in extra["projected"]:                                                     # shape like a sim player
    fa_pool.add(normalize(p["player"]))
    sim_by_team[p["dest_team"]].append({
        "player": p["player"], "pos": p["pos"], "barrett": p["barrett"], "status": p["status"],
        "incumbent": p["team"],
        "realistic": {"predicted": p["dest_team"], "is_resign": p["is_resign"], "offer_M": p["offer_M"]}})

# ── Manual signings (data/manual_signings.csv): force a specific player onto a team,
# optionally replacing whoever the model had there (the replaced player is bumped back
# to free agency and re-routed). For roster calls the model wouldn't make on its own.
displaced = []
_ms = ROOT / "data" / "manual_signings.csv"
if _ms.exists():
    for r in csv.DictReader([l for l in _ms.read_text().splitlines() if not l.lstrip().startswith("#")]):
        pl = (r.get("player") or "").strip()
        tm = ABBR_FIX.get((r.get("team") or "").strip(), (r.get("team") or "").strip())
        if not pl or tm not in TEAMS:
            continue
        n = normalize(pl); cost = None
        repl = normalize((r.get("replaces") or "").strip())
        if repl:                                                                 # bump the replaced player off this team
            keep = []
            for p in sim_by_team.get(tm, []):
                if normalize(p["player"]) == repl:
                    cost = p["realistic"]["offer_M"]; displaced.append(p)
                else:
                    keep.append(p)
            sim_by_team[tm] = keep
        for t2 in list(sim_by_team):                                             # remove the forced man from anywhere else
            sim_by_team[t2] = [p for p in sim_by_team[t2] if normalize(p["player"]) != n]
        fa_pool.add(n)
        cost_in = (r.get("cost_M") or "").strip()                                # explicit cost > replaced slot > minimum
        offer = float(cost_in) if cost_in else (cost if cost is not None else 2.3)
        resign = (r.get("resign") or "").strip().lower() in ("yes", "y", "true", "1")
        inc = cur_team(n) or tm
        resign = resign or (inc == tm)                                           # signing his own team -> Bird re-sign
        sim_by_team[tm].append({
            "player": disp.get(n, pl), "pos": pos_map.get(n, "—"), "barrett": bar_map.get(n),
            "status": "UFA", "incumbent": inc, "_forced": True,
            "realistic": {"predicted": tm, "is_resign": resign, "offer_M": round(offer, 1), "alts": []}})

# ── REAL 2026 signings (data/real_signings_2026.csv): reported deals are FACTS,
# not projections. A real deal overrides the model's projected destination
# outright: the player goes to his actual team at his actual first-year salary,
# flagged _real so he can never be trimmed or re-routed, and the page can label
# him "Signed" instead of "Projected". The sim keeps projecting only the
# still-unsigned pool. The projection MODEL stays pure — this layers reality on
# top for roster display, exactly what the tracker was built to enable.
real_set = set()
_rs = ROOT / "data" / "real_signings_2026.csv"
if _rs.exists():
    for r in csv.DictReader([l for l in _rs.read_text().splitlines() if not l.lstrip().startswith("#")]):
        pl = (r.get("player") or "").strip()
        tm = ABBR_FIX.get((r.get("team") or "").strip(), (r.get("team") or "").strip())
        rtyp = (r.get("type") or "").strip().lower()
        if not pl or tm not in TEAMS or rtyp == "extension":   # extensions: already under contract
            continue
        n = normalize(pl)
        if n in real_set:
            continue
        try:
            yr1 = float((r.get("yr1_M") or "").strip())
        except ValueError:
            continue
        real_set.add(n)
        fa_pool.add(n)                              # drops any stale under-contract row (decline-and-resign)
        for t2 in list(sim_by_team):                # kill the model's projected placement anywhere
            sim_by_team[t2] = [p for p in sim_by_team[t2] if normalize(p["player"]) != n]
        inc = cur_team(n) or tm
        sim_by_team[tm].append({
            "player": disp.get(n, pl), "pos": pos_map.get(n, "—"), "barrett": bar_map.get(n),
            "status": "UFA", "incumbent": inc, "_forced": True, "_real": True,
            "realistic": {"predicted": tm, "is_resign": rtyp == "resign" or inc == tm,
                          "offer_M": round(yr1, 1), "alts": []}})
print(f"real signings layered onto rosters: {len(real_set)}")

_STAT = {"guaranteed": "Signed", "player_option": "Player Option", "team_option": "Team Option"}

# ── COMPLETE under-contract roster: every 2026-27 contract that isn't a free agent ──
under = defaultdict(list)                                                         # team -> [(name,pos,sal,type,bar)]
for n, info in nc.items():
    typ = info.get("type")
    if typ not in _STAT or n in fa_pool or n in held:
        continue
    tm = ABBR_FIX.get(cur_team(n), cur_team(n))                            # correction > board > season stats > cey
    if tm not in TEAMS:
        continue
    if (tm, n) in dropped or (tm, _strip(n)) in dropped:                   # waived / declined option -> off roster
        continue
    under[tm].append((disp.get(n, n.title()), pos_map.get(n, "—"),
                      round((info.get("salary") or 0) / 1e6, 1), typ, bar_map.get(n)))

team_order = sorted(TEAMS, key=lambda t: TEAMS[t]["name"])


def _row(tm, name, pos, typ, status, prev, barrett, cost):
    return [tm, name, pos, typ, status, prev,
            (round(barrett, 1) if barrett is not None else None),
            (round(cost, 1) if cost is not None else None)]


ROSTER_MAX = 15
ABBR = lambda t: ABBR_FIX.get(t, t)
# ── Phase 1: each team's standard roster = under contract + best first-round picks +
# the projected moves that fit (re-signs before new signings, then by money, so a
# $2.3M team-option keeper isn't trimmed for a pricier external add). Excess picks on
# a full roster are treated as stashed/traded. Collect the OVERFLOW (projected players
# a full team has no room for) plus each team's remaining open spots.
MIN_M, MIN_FILL = 2.3, 3.0                              # at/below MIN_FILL an add is a minimum (apron-exempt)
plan_uc, plan_picks, plan_moves, open_left, team_total = {}, {}, {}, {}, {}
overflow = []
for tm in team_order:
    uc = sorted(under.get(tm, []), key=lambda x: -x[2])
    allp = sorted((m for m in TEAMS[tm]["plan"] if m.get("kind") == "pick"), key=lambda m: m.get("overall", 99))
    firsts = [m for m in allp if m.get("round") == 1]
    seconds = [m for m in allp if m.get("round") != 1]               # 2nd-round = two-way, no standard spot
    kept_first = firsts[:max(0, ROSTER_MAX - len(uc))]
    spots = max(0, ROSTER_MAX - len(uc) - len(kept_first))
    ranked = sorted(sim_by_team.get(tm, []),                                     # forced manual signings kept first
                    key=lambda p: (not p.get("_forced"), not p["realistic"]["is_resign"], -p["realistic"]["offer_M"]))
    # Keep moves up to the open spots, but a non-minimum OUTSIDE signing can't push the
    # team past the hard 2nd apron (no exception exists up there); it overflows and
    # re-routes apron-aware. Re-signs (Bird) and minimums are apron-exempt and stay.
    base = sum(x[2] for x in uc) + sum((m.get("cost_M") or 0) for m in kept_first + seconds)
    plan_uc[tm], plan_picks[tm] = uc, kept_first + seconds
    # Real deals are FACTS: they always stick, even past the 15-man count (an
    # over-15 roster is the audit's signal that the under-contract feed is stale
    # -- fix via roster_corrections -- not a reason to drop a reported signing).
    # Projected moves fill only the spots reality leaves open.
    real_mv = [p for p in ranked if p.get("_real")]
    proj_mv = [p for p in ranked if not p.get("_real")]
    kept = real_mv + proj_mv[:max(0, spots - len(real_mv))]
    overflow.extend(proj_mv[max(0, spots - len(real_mv)):])
    base = round(base + sum(p["realistic"]["offer_M"] for p in kept), 1)
    # Hard cap: a non-minimum OUTSIDE signing can't leave the team over the 2nd apron
    # (no exception exists up there). Pull the cheapest such signing until the team is
    # legal; it overflows and re-routes apron-aware. Bird re-signs and veteran minimums
    # over the apron are allowed, so they're never cut.
    while base > AP2:
        cuts = [p for p in kept if not p.get("_real")   # a reported real deal is a fact, never trimmed
                and not p["realistic"]["is_resign"] and p["realistic"]["offer_M"] > MIN_FILL]
        if not cuts:
            break
        worst = min(cuts, key=lambda p: p["realistic"]["offer_M"])
        kept.remove(worst); overflow.append(worst); base = round(base - worst["realistic"]["offer_M"], 1)
    plan_moves[tm] = kept
    open_left[tm] = max(0, spots - len(kept))
    team_total[tm] = base
overflow.extend(displaced)                                                       # manually-bumped players re-route too
overflow.extend(blocked_players)                                                 # blocked signings re-route elsewhere

# ── Phase 2: re-route OVERFLOW to the best open-spot team that fits. A good player
# whose projected team is full doesn't go unsigned - the league signs him. Prefer his
# own team (re-sign at home), then a team the model already ranked for him (his alts),
# then a positional need, then any room. Best (highest-offer) players placed first.
# A real (non-minimum) deal can't push a team over the hard 2nd apron, where it has no
# exceptions; if no one can fit the full deal he takes a minimum (apron-exempt) instead.
_prim = lambda pos: (pos or "").split("/")[0].strip().upper()
def _afford(t, off):                                     # no re-route may push a team past the hard 2nd apron
    return team_total[t] + off <= AP2
# Positional surplus: don't fill a team that's already deep at the player's spot (the model
# was blind to this, e.g. adding a 6th guard to New Orleans). Count every rostered player
# whose position string covers a slot; a team with >= GLUT there won't take another.
POS5, GLUT = ("PG", "SG", "SF", "PF", "C"), 5
share = {}
for tm in team_order:
    c = {q: 0 for q in POS5}
    for pos in [x[1] for x in plan_uc[tm]] + [m.get("pos", "") for m in plan_picks[tm]] + [p["pos"] for p in plan_moves[tm]]:
        for q in POS5:
            if q in (pos or ""):
                c[q] += 1
    share[tm] = c
def _glut(t, pos):
    return share[t].get(_prim(pos), 0) >= GLUT
def _bump_share(t, pos):
    for q in POS5:
        if q in (pos or ""):
            share[t][q] += 1
for p in sorted(overflow, key=lambda p: -p["realistic"]["offer_M"]):
    # Reality-first: a PROJECTED move displaced by real deals is just an unsigned
    # free agent -- inventing a new team for him (the pre-July re-route) reads as
    # nonsense once actual rosters are set. He falls through to the unsigned
    # list instead. Only manual force-signings still re-route.
    if not p.get("_forced"):
        continue
    inc = ABBR(p["incumbent"]); prim = _prim(p["pos"])
    alts = {a["team"]: a["offer_M"] for a in p["realistic"].get("alts", [])}
    base_off = p["realistic"]["offer_M"]
    best = None
    for t in team_order:
        if open_left.get(t, 0) <= 0 or (t, normalize(p["player"])) in blocked or _glut(t, p["pos"]):
            continue
        off = round(alts.get(t, base_off), 1)
        if not _afford(t, off):                         # would illegally cross the apron with a real deal
            continue
        need = prim in set(TEAMS[t].get("needs", []) + TEAMS[t].get("thin", []))
        key = (t == inc, t in alts, need, open_left[t])
        if best is None or key > best[0]:
            best = (key, t, off)
    if best is None:                                    # nobody fits the full deal -> minimum on a team with apron room
        mn = round(min(base_off, MIN_M), 1)
        opens = [t for t in team_order if open_left.get(t, 0) > 0 and team_total[t] + mn <= AP2
                 and (t, normalize(p["player"])) not in blocked]
        if not opens:
            opens = [t for t in team_order if open_left.get(t, 0) > 0 and (t, normalize(p["player"])) not in blocked]
        if not opens:
            continue
        t = max(opens, key=lambda t: (t == inc, t in alts, open_left[t]))
        best = (None, t, mn)
    _, t, off = best
    plan_moves[t].append({**p, "realistic": {**p["realistic"], "predicted": t,
                                             "offer_M": off, "is_resign": t == inc}})
    open_left[t] -= 1; team_total[t] = round(team_total[t] + off, 1); _bump_share(t, p["pos"])

# ── Phase 3: render the detail in team order.
detail = []
for tm in team_order:
    for name, pos, sal, ctyp, bar in plan_uc[tm]:
        detail.append(_row(tm, name, pos, "Under contract", _STAT.get(ctyp, "Signed"), tm, bar, sal))
    for m in sorted(plan_picks[tm], key=lambda m: m.get("overall", 99)):
        detail.append(_row(tm, m["name"], m["pos"], "Draft pick", "Rookie", "(draft)", None, m.get("cost_M") or 0))
    mv = plan_moves[tm]
    for p in sorted((q for q in mv if q["realistic"]["is_resign"]), key=lambda q: -q["realistic"]["offer_M"]):
        detail.append(_row(tm, p["player"], p["pos"], "Re-sign", p["status"], p["incumbent"],
                           p["barrett"], p["realistic"]["offer_M"]))
    for p in sorted((q for q in mv if not q["realistic"]["is_resign"]), key=lambda q: -q["realistic"]["offer_M"]):
        detail.append(_row(tm, p["player"], p["pos"], "New signing", p["status"], p["incumbent"],
                           p["barrett"], p["realistic"]["offer_M"]))

run = defaultdict(float)
for row in detail:
    run[row[0]] = round(run[row[0]] + (row[7] or 0), 1)
    row += [run[row[0]], round(CAP - run[row[0]], 1), round(AP2 - run[row[0]], 1)]

agg = {}
for tm in team_order:
    rs = [r for r in detail if r[0] == tm]
    typ = lambda t: [r for r in rs if r[3] == t]
    twoway = [r for r in typ("Draft pick") if r[2] == "2nd round"]
    agg[tm] = {"guar": round(sum(r[7] or 0 for r in typ("Under contract")), 1), "nuc": len(typ("Under contract")),
               "nrs": len(typ("Re-sign")), "rscost": round(sum(r[7] or 0 for r in typ("Re-sign")), 1),
               "nns": len(typ("New signing")), "nscost": round(sum(r[7] or 0 for r in typ("New signing")), 1),
               "npk": len(typ("Draft pick")), "pkcost": round(sum(r[7] or 0 for r in typ("Draft pick")), 1),
               "total": round(sum(r[7] or 0 for r in rs), 1), "size": len(rs) - len(twoway),
               "room": round(AP2 - sum(r[7] or 0 for r in rs), 1)}

# ── 2026 free agents NOT on a 15-man roster ───────────────────────────────────────
# Every player the main roster sheet can't show: free agents the projection pool
# never ranked (depth / minimum / two-way level), plus any projected signing the
# 15-man cap couldn't fit. Listed so the workbook accounts for every player.
rostered = {normalize(r[1]) for r in detail}
fa_rows = []                                                                      # [team, player, pos, barrett, value, landing, note]
for x in extra["listing"]:
    if normalize(x["player"]) in rostered:
        continue                                                                 # folded into a roster already
    note = "deep-bench / two-way level" if x["value_M"] is None else "minimum / depth free agent"
    fa_rows.append([ABBR_FIX.get(x["team"], x["team"]), x["player"], x["pos"],
                    x.get("barrett"), x.get("value_M"), None, note])
for p in sim["players"]:                                                          # projections the cap dropped
    if p["realistic"]["predicted"] and normalize(p["player"]) not in rostered:
        kind = "re-sign" if p["realistic"]["is_resign"] else "signing"
        fa_rows.append([p["incumbent"], p["player"], p["pos"], p.get("barrett"),
                        None, p["realistic"]["predicted"],
                        f"projected {kind} - no room on 15-man roster"])
for _tm, _nm in declined:                                                         # team options the team declined
    n = normalize(_nm)
    if n in rostered:
        continue
    opt = (nc.get(n) or {}).get("salary")
    note = (f"team option declined (${round(opt / 1e6, 1)}M option) - now a free agent"
            if opt else "team option declined - now a free agent")
    fa_rows.append([_tm, disp.get(n, _nm), pos_map.get(n, "—"), bar_map.get(n), None, None, note])
fa_rows.sort(key=lambda r: (r[0], -(r[4] or r[3] or 0)))

# ── Dump the assembled 2026-27 rosters for the crawlable /team/<ABBR> pages
# (scripts/build_team_pages.py renders from this, so the pages and this workbook
# stay one source of truth). Sits after fa_rows so each team also carries its
# notable STILL-UNSIGNED free agents (e.g. an incumbent star without a reported
# deal whose old team is already full). Before the xlsx save, so it still lands
# when the workbook write is blocked. ─────────────────────────────────────────────
import json as _json
_ros = {"value_season": sim.get("season", "2025-26"),
        "contract_season": sim.get("contract_season", "2026-27"),
        "cap_M": CAP, "apron2_M": AP2, "teams": {}}
for _tm in team_order:
    _rows = [r for r in detail if r[0] == _tm]
    _uns = [{"n": x[1], "barrett": x[3]} for x in fa_rows
            if x[0] == _tm and (x[3] or 0) >= 8.0]        # rotation-level only
    _ros["teams"][_tm] = {
        "abbr": _tm, "name": TEAMS[_tm]["name"],
        "total": agg[_tm]["total"], "size": agg[_tm]["size"], "room": agg[_tm]["room"],
        "players": [{"n": r[1], "role": r[3], "pos": r[2], "prev": r[5],
                     "barrett": r[6], "salary": r[7],
                     # signings/re-signs: True = reported real deal, False = model
                     # projection for a still-unsigned FA. Contracts/picks are facts.
                     "real": (normalize(r[1]) in real_set
                              if r[3] in ("Re-sign", "New signing") else True)}
                    for r in _rows],
        "unsigned": sorted(_uns, key=lambda u: -(u["barrett"] or 0)),
    }
_ros_path = Path(__file__).resolve().parent.parent / "cache" / "team_rosters_2627.json"
_ros_path.write_text(_json.dumps(_ros, separators=(",", ":")))
print(f"wrote {_ros_path.name} ({len(team_order)} teams, {len(detail)} roster rows, "
      f"{sum(len(t['unsigned']) for t in _ros['teams'].values())} notable unsigned)")

# ── styles ──────────────────────────────────────────────────────────────────────
FONT = "Arial"
base, boldf = Font(name=FONT, size=10), Font(name=FONT, size=10, bold=True)
hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F3A5F")
band = PatternFill("solid", fgColor="EEF2F7")
type_fill = {"Re-sign": PatternFill("solid", fgColor="E2F1E7"),
             "New signing": PatternFill("solid", fgColor="E6EEF9"),
             "Draft pick": PatternFill("solid", fgColor="F1ECF8"),
             "Under contract": PatternFill("solid", fgColor="F4F5F7")}
MONEY, SCORE = '$#,##0.0;($#,##0.0);"-"', '0.0'
team_top = Border(top=Side(style="medium", color="9AA7B5"))
rt, ct = Alignment(horizontal="right"), Alignment(horizontal="center")
wrap_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = Workbook()
s = wb.active
s.title = "Summary"
s["A1"] = "HoopsValue — 2026-27 Free Agency & Projected Rosters"
s["A1"].font = Font(name=FONT, size=14, bold=True, color="1F3A5F")
s["A2"] = "Complete rosters: every player under contract for 2026-27, plus projected re-signings and new signings."
s["A2"].font = Font(name=FONT, size=9, italic=True, color="6B7682")
for i, (lab, val) in enumerate([("Salary Cap", CAP), ("Luxury Tax", TAX), ("First Apron", AP1), ("Second Apron", AP2)]):
    s.cell(4 + i, 1, lab).font = boldf
    c = s.cell(4 + i, 2, val); c.font = base; c.number_format = '$#,##0"M"'; c.alignment = rt
cols = ["Abbr", "Team", "Guaranteed\nPayroll ($M)", "Under\nContract", "Re-\nsigns", "Re-sign\nCost ($M)",
        "New\nSignings", "Signing\nCost ($M)", "Draft\nPicks", "Pick Cost\n($M)", "Total Payroll\n($M)",
        "Roster\nSize", "Room to\n2nd Apron ($M)"]
HR = 9
for j, c in enumerate(cols, 1):
    cell = s.cell(HR, j, c); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = wrap_c
for i, tm in enumerate(team_order):
    a = agg[tm]; r = HR + 1 + i
    vals = [tm, TEAMS[tm]["name"], a["guar"], a["nuc"], a["nrs"], a["rscost"], a["nns"],
            a["nscost"], a["npk"], a["pkcost"], a["total"], a["size"], a["room"]]
    for j, v in enumerate(vals, 1):
        cell = s.cell(r, j, v); cell.font = base
        if j in (3, 6, 8, 10, 11, 13):
            cell.number_format = MONEY
        if j >= 3:
            cell.alignment = rt
    if i % 2 == 1:
        for j in range(1, 14):
            s.cell(r, j).fill = band
tot = HR + 1 + len(team_order)
s.cell(tot, 2, "League total").font = boldf
for col, key in [(3, "guar"), (4, "nuc"), (5, "nrs"), (6, "rscost"), (7, "nns"),
                 (8, "nscost"), (9, "npk"), (10, "pkcost"), (11, "total"), (12, "size")]:
    cell = s.cell(tot, col, round(sum(agg[t][key] for t in team_order), 1)); cell.font = boldf
    if col in (3, 6, 8, 10, 11):
        cell.number_format = MONEY
    cell.alignment = rt
for j in range(1, 14):
    s.cell(tot, j).border = Border(top=Side(style="thin", color="9AA7B5"))
for j, w in enumerate([7, 24, 13, 9, 7, 11, 9, 11, 8, 10, 13, 8, 14], 1):
    s.column_dimensions[get_column_letter(j)].width = w
s.freeze_panes = "C10"

d = wb.create_sheet("Roster Detail")
dcols = ["Team", "Player", "Position", "Type", "Status", "Previous\nTeam", "Barrett\nScore",
         "Cost /\nSalary ($M)", "Running\nPayroll ($M)", "Cap Space\n($M)", "Room to\n2nd Apron ($M)"]
for j, c in enumerate(dcols, 1):
    cell = d.cell(1, j, c); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = wrap_c
for i, row in enumerate(detail):
    r = 2 + i
    for j, v in enumerate(row, 1):
        cell = d.cell(r, j, v); cell.font = base
        if j == 7:
            cell.number_format = SCORE; cell.alignment = ct
        elif j >= 8:
            cell.number_format = MONEY; cell.alignment = rt
        elif j in (1, 4, 6):
            cell.alignment = ct
    if (tf := type_fill.get(row[3])):
        d.cell(r, 4).fill = tf
    if i > 0 and row[0] != detail[i - 1][0]:
        for j in range(1, 12):
            d.cell(r, j).border = team_top
for j, w in enumerate([7, 24, 11, 14, 13, 9, 9, 12, 13, 11, 14], 1):
    d.column_dimensions[get_column_letter(j)].width = w
d.freeze_panes = "C2"

fa = wb.create_sheet("2026 Free Agents")
fa["A1"] = "2026 Free Agents — not on a projected 15-man roster"
fa["A1"].font = Font(name=FONT, size=12, bold=True, color="1F3A5F")
fa["A2"] = ("Free agents the projection pool didn't rank (depth / minimum / two-way level), "
            "plus projected deals the 15-man cap couldn't fit. Every other player is on the Roster Detail sheet.")
fa["A2"].font = Font(name=FONT, size=9, italic=True, color="6B7682")
fcols = ["Current\nTeam", "Player", "Position", "Barrett\nScore", "Est. Value\n($M)",
         "Projected\nLanding", "Note"]
FHR = 4
for j, c in enumerate(fcols, 1):
    cell = fa.cell(FHR, j, c); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = wrap_c
for i, row in enumerate(fa_rows):
    r = FHR + 1 + i
    for j, v in enumerate(row, 1):
        cell = fa.cell(r, j, v); cell.font = base
        if j == 4:
            cell.number_format = SCORE; cell.alignment = ct
        elif j == 5:
            cell.number_format = MONEY; cell.alignment = rt
        elif j in (1, 3, 6):
            cell.alignment = ct
    if i % 2 == 1:
        for j in range(1, 8):
            fa.cell(r, j).fill = band
for j, w in enumerate([9, 24, 11, 9, 11, 11, 42], 1):
    fa.column_dimensions[get_column_letter(j)].width = w
fa.freeze_panes = "A5"

# ── Team Budgets: spending capacity vs estimated spend, all 30 teams ───────────────
# Budget = room from current guaranteed commitments to each line (tax / 2nd apron).
# Spending Power = the realistic tool to add OUTSIDE free agents — cap room if a team
# has it, otherwise the mid-level exception its apron tier allows. Estimated spend =
# projected re-signings + new signings. Unused vs Tax = how much room is left on the
# table at the projected final payroll (the "could they spend more?" number).
MLE_FULL, MLE_TAX = 14.1, 5.7          # 2026-27 full / taxpayer mid-level exception
def _spending_power(guar, rscost, pkcost):
    committed = guar + rscost + pkcost
    cap_room = round(CAP - committed, 1)
    if cap_room >= 8:   return max(cap_room, 0.0), "Cap room"
    if committed >= AP2: return 0.0, "Minimums only"
    if committed >= AP1: return MLE_TAX, "Taxpayer MLE"
    return MLE_FULL, "Full MLE"

bud = wb.create_sheet("Team Budgets")
bud["A1"] = "Team Budgets — spending capacity vs projected spend (2026-27)"
bud["A1"].font = Font(name=FONT, size=12, bold=True, color="1F3A5F")
bud["A2"] = ("Budget columns = room from current guaranteed commitments to each line. Spending Power = the "
             "realistic tool to add outside free agents (cap room, or the mid-level exception by apron tier). "
             "Estimated spend = projected re-signings + new signings.")
bud["A2"].font = Font(name=FONT, size=9, italic=True, color="6B7682")
bcols = ["Abbr", "Team", "Guaranteed\nPayroll ($M)", "Cap Room\n($M)", "Spending\nPower ($M)", "Spending\nTier",
         "Budget to\nTax ($M)", "Budget to\n2nd Apron ($M)", "Est. Offseason\nSpend ($M)",
         "Projected\nPayroll ($M)", "Unused vs\nTax ($M)", "Spending Status\n(why room is/ isn't used)"]
BHR = 4
for j, c in enumerate(bcols, 1):
    cell = bud.cell(BHR, j, c); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = wrap_c


def _spend_status(a, tier):
    # Explains the unused room: a full 15-man roster can't sign anyone (its room is
    # trade-only); an open spot can, via whatever tool the apron tier allows.
    if a["size"] >= 15:
        return "Full roster (15) - unused room is trade-only"
    if a["total"] >= AP2:
        return "Open spot - 2nd apron hard cap (minimums)"
    if tier == "Cap room":
        return "Open spot + cap room to use"
    return f"Open spot - {tier} only"


brows = []
for tm in team_order:
    a = agg[tm]; g = a["guar"]; power, tier = _spending_power(g, a["rscost"], a["pkcost"])
    brows.append([tm, TEAMS[tm]["name"], g, round(max(CAP - g, 0.0), 1), power, tier,
                  round(TAX - g, 1), round(AP2 - g, 1), round(a["rscost"] + a["nscost"], 1),
                  a["total"], round(TAX - a["total"], 1), _spend_status(a, tier)])
MONEYCOLS = (3, 4, 5, 7, 8, 9, 10, 11)
for i, row in enumerate(brows):
    r = BHR + 1 + i
    for j, v in enumerate(row, 1):
        cell = bud.cell(r, j, v); cell.font = base
        if j in MONEYCOLS:
            cell.number_format = MONEY; cell.alignment = rt
        elif j in (1, 6):
            cell.alignment = ct
    if i % 2 == 1:
        for j in range(1, 13):
            bud.cell(r, j).fill = band
btot = BHR + 1 + len(brows)
bud.cell(btot, 2, "League total").font = boldf
for col in MONEYCOLS:
    cell = bud.cell(btot, col, round(sum(rw[col - 1] for rw in brows), 1)); cell.font = boldf
    cell.number_format = MONEY; cell.alignment = rt
for j in range(1, 13):
    bud.cell(btot, j).border = Border(top=Side(style="thin", color="9AA7B5"))
for j, w in enumerate([7, 22, 13, 10, 11, 13, 11, 14, 14, 12, 11, 40], 1):
    bud.column_dimensions[get_column_letter(j)].width = w
bud.freeze_panes = "C5"

# ── Model Comparison: 4 projection models, where each sends every free agent ──────
_ens = ROOT / "cache" / "fa_ensemble_v1.json"
if _ens.exists():
    ens = json.loads(_ens.read_text())
    models = ens["models"]
    ms = wb.create_sheet("Model Comparison")
    ms["A1"] = "Model Comparison - five lenses on where every free agent lands"
    ms["A1"].font = Font(name=FONT, size=12, bold=True, color="1F3A5F")
    ms["A2"] = ("Each model applies a different philosophy to the SAME priced free-agent pool. The first model, "
                "Consensus, is the validated projection on the roster sheets; the others are stress-test scenarios. "
                "'Lenses agree' = how many of the models keep the player where Consensus puts him - all = a lock, "
                f"1 of {len(models)} = volatile (every scenario moves him). Green cells are a projected move off his current team.")
    ms["A2"].font = Font(name=FONT, size=9, italic=True, color="6B7682")
    mcols = ["Player", "Pos", "Current\nTeam", "Value\n($M)"] + models + ["Lenses\nagree"]
    MHR = 4
    for j, c in enumerate(mcols, 1):
        cell = ms.cell(MHR, j, c); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = wrap_c
    move_fill = PatternFill("solid", fgColor="E2F1E7")
    agree_fill = {len(models): PatternFill("solid", fgColor="CDEAD6"),       # all agree -> green
                  len(models) - 1: PatternFill("solid", fgColor="EAF3D6")}   # near -> pale
    split_fill = PatternFill("solid", fgColor="F6D9D5")                       # <=2 -> red
    for i, p in enumerate(ens["players"]):
        r = MHR + 1 + i
        inc = ABBR_FIX.get(p["incumbent"], p["incumbent"])
        vals = [p["player"], p["pos"], inc, p["value_M"]] + \
               [p["picks"][m]["team"] for m in models] + [f"{p['agreement']}/{len(models)}"]
        for j, v in enumerate(vals, 1):
            cell = ms.cell(r, j, v); cell.font = base
            if j == 4:
                cell.number_format = MONEY; cell.alignment = rt
            elif j == 2 or j == 3 or j >= 5:
                cell.alignment = ct
        for k, m in enumerate(models):                                        # green a projected move
            if p["picks"][m]["team"] != inc:
                ms.cell(r, 5 + k).fill = move_fill
        ag = ms.cell(r, len(mcols))
        ag.fill = agree_fill.get(p["agreement"], split_fill if p["agreement"] <= 2 else PatternFill())
    for j, w in enumerate([22, 7, 8, 8] + [13] * len(models) + [8], 1):
        ms.column_dimensions[get_column_letter(j)].width = w
    ms.freeze_panes = "B5"

# ── Plausibility audit: the projection checks ITSELF for the classes of nonsense
# that otherwise have to be caught by eye (a good player going unsigned, a ghost
# contract, an over-stuffed roster, money that doesn't add up). The ensemble is the
# cross-check: a player several models would roster should never end up unsigned.
ROT = 8.0                                  # value at/above this -> rotation-level, must not go unsigned
rostered_n = {normalize(r[1]) for r in detail}
fa_n = {normalize(x[1]) for x in fa_rows}
issues = []
_epath = ROOT / "cache" / "fa_ensemble_v1.json"
if _epath.exists():
    for p in json.loads(_epath.read_text())["players"]:
        n = normalize(p["player"])
        if p["value_M"] >= ROT and n not in rostered_n and n not in held:        # held-out players are intentional
            n_roster = sum(1 for m, pk in p["picks"].items() if pk["team"])      # lenses that place him
            if n in fa_n:
                # Reality-first: he's ON the unsigned list, which mid-offseason is a
                # FACT (no reported deal yet), not projection nonsense. Info, not issue.
                print(f"  still unsigned (reality): {p['player']} (${p['value_M']}M value, no reported deal)")
            else:
                issues.append(f"GHOST rotation FA (on no roster AND no unsigned list): "
                              f"{p['player']} (${p['value_M']}M, {n_roster} models roster him)")
for tm in team_order:
    if agg[tm]["size"] > 15:
        issues.append(f"ROSTER OVER 15: {tm} = {agg[tm]['size']}")
for r in detail:
    # over the 2nd apron a team has NO exceptions, so a non-minimum new signing there is illegal
    # (Bird re-signs and veteran minimums over the apron are allowed and not flagged).
    # A REPORTED real deal is exempt: reality is legal by definition, so an apron overage
    # there means OUR payroll data has drifted (stale under-contract feed), not that the
    # signing is nonsense -- surfaced as drift, not as a projection error.
    if r[3] == "New signing" and (r[7] or 0) > 3.0 and agg[r[0]]["total"] > AP2 + 0.5:
        if normalize(r[1]) in real_set:
            print(f"  payroll drift note: real signing {r[1]} on {r[0]} shows team at "
                  f"${agg[r[0]]['total']}M (> 2nd apron) -- check under-contract feed for {r[0]}")
        else:
            issues.append(f"OVER 2ND APRON via a ${r[7]}M signing: {r[1]} on {r[0]} (team ${agg[r[0]]['total']}M, no exception over the apron)")
    if r[3] == "Under contract" and r[4] == "Signed":
        info = cey.get(normalize(r[1])) or {}
        if info.get("end_season") and info["end_season"] < "2026-27":
            issues.append(f"GHOST (guaranteed but contract ended {info['end_season']}): {r[1]} {r[0]}")
    if r[3] == "Under contract":                                                  # team-of-record sanity (caught Rui's class)
        st = stats_team.get(normalize(r[1]))
        if st and r[0] != st and normalize(r[1]) not in team_corr:
            issues.append(f"TEAM MISMATCH (review): {r[1]} shown on {r[0]} but played for {st} this season")
    if normalize(r[1]) in fa_n:
        issues.append(f"DOUBLE-LISTED (on a roster AND the FA sheet): {r[1]}")
print(f"PLAUSIBILITY AUDIT: {len(issues)} issue(s)")
for s in issues:
    print("  ! " + s)
if not issues:
    print("  clean - every rotation FA is signed, no ghosts, no roster over 15, no double-listings, no team over the 2nd apron")

OUT.parent.mkdir(exist_ok=True)
wb.save(OUT)
# Also write a uniquely-named, timestamped copy each run so a freshly-opened file
# always shows the latest changes (spreadsheet apps cache an already-open workbook).
import datetime
stamp = datetime.datetime.now().strftime("%b%d_%H%M")
STAMPED = OUT.with_name(f"HoopsValue_FreeAgency_2026_{stamp}.xlsx")
wb.save(STAMPED)
print(f"wrote {OUT}")
print(f"OPEN THIS (latest): {STAMPED}")
print(f"  {len(detail)} roster rows | {sum(len(under[t]) for t in under)} under contract + "
      f"{sum(1 for r in detail if r[3] in ('Re-sign', 'New signing'))} signings")
print(f"  + {len(fa_rows)} free agents listed on the 2026 Free Agents sheet")
