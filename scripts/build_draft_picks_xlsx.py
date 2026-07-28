"""Build the Draft Picks workbook from cache/pick_ledger.json.

data/pick_ledger.csv is the maintained source of truth (edit rows as trades
happen, re-run scripts/build_pick_ledger.py, then this). Output workbook:
  Sheet 1  Second Rounders - every 2027-2033 second, one row per pick
  Sheet 2  First Rounders  - every 2027-2033 first with protections/swaps
  Sheet 3  By Team         - per-team counts of controlled firsts/seconds
  Sheet 4  About           - as-of date, sources, how to maintain

Run: python3 scripts/build_draft_picks_xlsx.py [out.xlsx]
Default output: exports/HoopsValue_Draft_Picks.xlsx (NOT committed; the CSV is).
"""
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "cache" / "pick_ledger.json"
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "exports" / "HoopsValue_Draft_Picks.xlsx"

FONT = "Arial"
base = Font(name=FONT, size=10)
bold = Font(name=FONT, size=10, bold=True)
hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F3A5F")
band = PatternFill("solid", fgColor="EEF2F7")
team_fill = PatternFill("solid", fgColor="DCE6F2")
traded_font = Font(name=FONT, size=10, bold=True, color="8A2A1A")


def pick_sheet(wb, title, picks):
    ws = wb.create_sheet(title)
    heads = ["Origin", "Year", "Controlled by", "Traded?", "Protection / swap", "Notes"]
    ws.append(heads)
    for c in range(1, len(heads) + 1):
        ws.cell(1, c).font = hdr_font
        ws.cell(1, c).fill = hdr_fill
    picks = sorted(picks, key=lambda p: (p["origin"], p["year"]))
    prev = None
    for i, p in enumerate(picks, start=2):
        enc = p.get("protection", "") or (
            f"swap with {p['swap_with']}" if p.get("swap_with") else "")
        traded = p["controlled_by"] != p["origin"]
        ws.append([p["origin"], p["year"], p["controlled_by"],
                   "TRADED" if traded else "", enc, p.get("notes", "")])
        first = p["origin"] != prev
        prev = p["origin"]
        for c in range(1, len(heads) + 1):
            cell = ws.cell(i, c)
            cell.font = base
            if first:
                cell.fill = team_fill
            elif i % 2 == 0:
                cell.fill = band
        if first:
            ws.cell(i, 1).font = bold
        if traded:
            ws.cell(i, 3).font = traded_font
            ws.cell(i, 4).font = traded_font
    for j, w in enumerate([8, 7, 14, 10, 34, 60], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(picks) + 1}"


def main() -> None:
    data = json.loads(SRC.read_text())
    all_picks = [p for d in data["teams"].values() for p in d["controls"]]
    seconds = [p for p in all_picks if p.get("round", 1) == 2]
    firsts = [p for p in all_picks if p.get("round", 1) == 1]

    wb = Workbook()
    wb.remove(wb.active)
    pick_sheet(wb, "Second Rounders", seconds)
    pick_sheet(wb, "First Rounders", firsts)

    ts = wb.create_sheet("By Team")
    ts.append(["Team", "Firsts controlled", "Seconds controlled",
               "Own seconds kept", "Seconds acquired", "Own seconds traded away"])
    for c in range(1, 7):
        ts.cell(1, c).font = hdr_font
        ts.cell(1, c).fill = hdr_fill
    own_traded = defaultdict(int)
    for p in seconds:
        if p["controlled_by"] != p["origin"]:
            own_traded[p["origin"]] += 1
    for i, t in enumerate(sorted(data["teams"]), start=2):
        ctrl = data["teams"][t]["controls"]
        s = [p for p in ctrl if p.get("round", 1) == 2]
        ts.append([t, sum(1 for p in ctrl if p.get("round", 1) == 1), len(s),
                   sum(1 for p in s if p["origin"] == t),
                   sum(1 for p in s if p["origin"] != t), own_traded[t]])
        for c in range(1, 7):
            ts.cell(i, c).font = base
            if i % 2 == 0:
                ts.cell(i, c).fill = band
        ts.cell(i, 1).font = bold
    for j, w in enumerate([7, 16, 18, 16, 16, 20], 1):
        ts.column_dimensions[get_column_letter(j)].width = w
    ts.freeze_panes = "A2"

    ab = wb.create_sheet("About")
    lines = [
        ("HoopsValue Draft Picks Master List", bold),
        (f"As of: {data.get('asof', '')}", base),
        ("", base),
        ("Every first- and second-round pick, drafts 2027-2033, all 30 teams.", base),
        ("Second-round ownership verified league-wide by a both-sides research", base),
        ("workflow (RealGM future-picks pages cross-checked against Tankathon,", base),
        ("Spotrac and July 2026 trade reporting; every traded pick confirmed", base),
        ("from the origin side and the receiving side).", base),
        ("", base),
        ("To maintain: edit data/pick_ledger.csv as trades happen, re-run", base),
        ("scripts/build_pick_ledger.py, then this script. The ledger also", base),
        ("powers the Trade Machine's pick cards and Stepien-rule checks.", base),
    ]
    for i, (txt, f) in enumerate(lines, start=1):
        ab.cell(i, 1, txt).font = f
    ab.column_dimensions["A"].width = 72

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    n2 = sum(1 for p in seconds if p["controlled_by"] != p["origin"])
    print(f"wrote {OUT} ({len(firsts)} firsts, {len(seconds)} seconds, "
          f"{n2} seconds traded)")


if __name__ == "__main__":
    main()
