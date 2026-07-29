#!/usr/bin/env python3
"""Basketball Reference "Position Estimate" percentages for a list of players.

BBRef's play-by-play table carries the share of a player's minutes spent at each
of the five spots (PG% SG% SF% PF% C%), per season plus a career row. That is
the closest thing to ground truth for "what does this guy actually play", so it
is the right check on a hand-assigned position.

Writes exports/HoopsValue_Position_Estimates.xlsx: one row per player with the
career line, the 2025-26 line, the top two spots implied by each, and the
position currently on the master roster.

BBRef rate-limits hard (20 requests a minute, and it blocks for an hour if it
decides you are a bot), so this crawls slowly and caches every page it fetches
to cache/bref_pbp/ - a re-run costs nothing.

    python scripts/fetch_bref_position_pct.py
"""
import json
import re
import sys
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
from utils import normalize  # noqa: E402

SEASON_END = 2026                       # 2025-26
SEASON_LABEL = "2025-26"
OUT = ROOT / "exports" / "HoopsValue_Position_Estimates.xlsx"
CACHE = ROOT / "cache" / "bref_pbp"
UA = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"}
SPOTS = ["PG", "SG", "SF", "PF", "C"]
PAUSE = 8.0                             # well under their 20-a-minute ceiling

# the players repositioned by hand this session, and what they were set to,
# plus the two whose master-roster position still disagrees with the override
PLAYERS = {
    "Quentin Grimes": "SF/SG",
    "Matisse Thybulle": "SG/SF",
    "Ziaire Williams": "PF/SF",
    "Derrick White": "SG/PG",
    "OG Anunoby": "PF/SF",
    "Kawhi Leonard": "SF/PF",
    "Tyler Herro": "SG/PG",
    "Kyle Kuzma": "PF/SF",
    "Paolo Banchero": "PF/SF",
    "Luguentz Dort": "SF/PF",
    "Kobe Sanders": "SG",
}


def _get(url: str, key: str) -> str:
    """Fetch with an on-disk cache, so a re-run never re-hits BBRef."""
    CACHE.mkdir(parents=True, exist_ok=True)
    p = CACHE / f"{key}.html"
    if p.exists() and p.stat().st_size > 10_000:
        return p.read_text(encoding="utf-8")
    r = requests.get(url, headers=UA, timeout=30)
    if r.status_code == 429:
        raise RuntimeError("basketball-reference is rate limiting (429); wait and re-run")
    r.raise_for_status()
    # BBRef serves UTF-8 without a charset header, and requests then falls back
    # to ISO-8859-1 for text/*, which turns "Jokic" with its accent into
    # mojibake and drops the player from every name lookup
    r.encoding = "utf-8"
    p.write_text(r.text, encoding="utf-8")
    time.sleep(PAUSE)
    return r.text


def player_ids(end_year: int) -> dict:
    """{normalized name: bbref id} from a season's per-game page."""
    html = _get(f"https://www.basketball-reference.com/leagues/NBA_{end_year}_per_game.html",
                f"per_game_{end_year}")
    ids = {}
    for m in re.finditer(r'/players/./([a-z0-9]+)\.html"[^>]*>([^<]+)</a>', html):
        ids.setdefault(normalize(m.group(2)), m.group(1))
    return ids


def season_pcts(end_year: int) -> dict:
    """{normalized name: {spot: pct, mp, team}} for one season, in ONE request.

    BBRef publishes the play-by-play table league-wide, so the whole season
    costs a single page instead of one per player (or per team). The per-team
    pages are the fallback if that page ever moves.
    """
    out = {}
    try:
        html = _get(f"https://www.basketball-reference.com/leagues/"
                    f"NBA_{end_year}_play-by-play.html", f"league_pbp_{end_year}")
        sources = [("league", html)]
    except Exception as e:
        print(f"  league play-by-play page unavailable ({e}); falling back to team pages")
        sources = []
        for ab in TEAMS_BREF:
            try:
                sources.append((ab, _get(
                    f"https://www.basketball-reference.com/teams/{ab}/{end_year}"
                    f"/play-by-play/", f"team_pbp_{ab}_{end_year}")))
            except Exception as te:
                print(f"    {ab}: {te}")
    for _src, html in sources:
        for chunk in [html] + re.findall(r"<!--(.*?)-->", html, re.S):
            if "pbp" not in chunk:
                continue
            soup = BeautifulSoup(chunk, "html.parser")
            tbl = soup.find("table", id=re.compile(r"pbp"))
            if tbl is None:
                continue
            keys = _pct_keys(tbl)
            if len(keys) < 5:
                continue
            for tr in tbl.find_all("tr"):
                cells = {c.get("data-stat"): c.get_text(strip=True)
                         for c in tr.find_all(["th", "td"]) if c.get("data-stat")}
                nm = cells.get("name_display") or cells.get("player")
                if not nm or nm == "Player":
                    continue
                rec = pct(cells, keys)
                rec["mp"] = cells.get("mp", "")
                rec["team"] = (cells.get("team_name_abbr") or cells.get("team_id") or "")
                prev = out.get(normalize(nm))
                # a traded player has one row per team plus a combined row;
                # keep whichever line covers the most minutes
                if prev is None or _num(rec["mp"]) > _num(prev.get("mp")):
                    out[normalize(nm)] = rec
            break
    return out


def _num(v) -> float:
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


TEAMS_BREF = ["ATL", "BOS", "BRK", "CHO", "CHI", "CLE", "DAL", "DEN", "DET", "GSW",
              "HOU", "IND", "LAC", "LAL", "MEM", "MIA", "MIL", "MIN", "NOP", "NYK",
              "OKC", "ORL", "PHI", "PHO", "POR", "SAC", "SAS", "TOR", "UTA", "WAS"]


def _pct_keys(table) -> dict:
    """{spot: data-stat} read off the header, so we do not hard-code BBRef's
    column ids (they have changed them before)."""
    keys = {}
    for th in table.find_all("th"):
        txt = th.get_text(strip=True).replace("%", "").upper()
        stat = th.get("data-stat")
        if txt in SPOTS and stat and stat not in keys.values():
            keys[txt] = stat
    return keys


def pbp_rows(pid: str):
    """(rows, {spot: data-stat}) from a player's play-by-play table.

    BBRef ships every table after the first inside an HTML comment, so the
    parser has to look in the comments too or the table is simply not there.
    """
    html = _get(f"https://www.basketball-reference.com/players/{pid[0]}/{pid}.html", pid)
    for chunk in [html] + re.findall(r"<!--(.*?)-->", html, re.S):
        if "pbp" not in chunk:
            continue
        soup = BeautifulSoup(chunk, "html.parser")
        tbl = soup.find("table", id=re.compile(r"pbp"))
        if tbl is None:
            continue
        keys = _pct_keys(tbl)
        if len(keys) < 5:
            continue
        rows = []
        for tr in tbl.find_all("tr"):
            cells = {c.get("data-stat"): c.get_text(strip=True)
                     for c in tr.find_all(["th", "td"]) if c.get("data-stat")}
            if cells:
                rows.append(cells)
        return rows, keys
    return [], {}


def pct(row: dict, keys: dict) -> dict:
    out = {}
    for s in SPOTS:
        v = row.get(keys.get(s, ""), "")
        try:
            out[s] = float(str(v).replace("%", "")) if v else 0.0
        except ValueError:
            out[s] = 0.0
    return out


def top_two(p: dict) -> str:
    """The two spots with the most minutes, primary first. One spot only when
    the second is negligible: a player at 96 percent SF is not a SF/PF."""
    if not p or sum(p.get(s, 0) or 0 for s in SPOTS) == 0:
        return ""              # the dict also carries mp/team, so sum SPOTS only
    order = sorted(SPOTS, key=lambda s: -p.get(s, 0.0))
    return order[0] if p.get(order[1], 0) < 10 else f"{order[0]}/{order[1]}"


def _label(row: dict) -> str:
    for k in ("year_id", "season", "ranker"):
        if row.get(k):
            return row[k]
    return ""


FIRST_SEASON_END = 2004                 # LeBron's rookie year; nobody active is older
AUDIT = []                              # every rostered player, filled by audit()


def career_pcts(first_end: int, last_end: int):
    """Career percentages for EVERY player, from one league page per season.

    BBRef's own career row is a minutes-weighted average of the seasons, so
    aggregating the season pages reproduces it: 23 requests for the whole
    league instead of one player page each for 440 players. Returns
    ({name: {spot: pct, mp}}, {name: [(season_end, pcts), ...]}) so callers can
    also ask "the most recent season he actually played enough of".
    """
    tot, hist = {}, {}
    for end in range(first_end, last_end + 1):
        try:
            rows = season_pcts(end)
        except Exception as e:
            print(f"  {end - 1}-{str(end)[2:]}: {e}")
            continue
        for name, rec in rows.items():
            mp = _num(rec.get("mp"))
            if mp <= 0:
                continue
            acc = tot.setdefault(name, {s: 0.0 for s in SPOTS} | {"mp": 0.0})
            for s in SPOTS:
                acc[s] += (rec.get(s) or 0.0) * mp
            acc["mp"] += mp
            hist.setdefault(name, []).append((end, rec))
        print(f"  {end - 1}-{str(end)[2:]}: {len(rows)} players")
    out = {}
    for name, acc in tot.items():
        mp = acc["mp"]
        out[name] = {s: round(acc[s] / mp, 1) for s in SPOTS} | {"mp": int(mp)}
    for h in hist.values():
        h.sort(key=lambda x: -x[0])
    return out, hist


def recent_qualifying(hist_rows, floor: int = 500):
    """The most recent season the player cleared the minutes floor, so a
    5-game season does not decide what position he plays. Falls back to his
    most recent season when he never cleared it."""
    for end, rec in hist_rows or []:
        if _num(rec.get("mp")) >= floor:
            return end, rec
    return (hist_rows or [(None, None)])[0]


def audit(cur_all: dict) -> list:
    """Same comparison for every player on the 2026-27 master roster."""
    import csv
    p = ROOT / "data" / "master_roster.csv"
    rows = []
    lines = [l for l in p.read_text().splitlines() if not l.lstrip().startswith("#")]
    for r in csv.DictReader(lines):
        nm = (r.get("player") or "").strip()
        if not nm or (r.get("kind") or "") == "free_agent":
            continue
        cur = cur_all.get(normalize(nm))
        if not cur:
            continue                    # did not play last season, nothing to check
        master = (r.get("pos") or "").strip()
        top = top_two(cur)
        rec = {"Player": nm, "Team": (r.get("team") or "").strip(), "Master pos": master,
               f"{SEASON_LABEL} top 2": top, "Agrees": "yes" if top == master else "no",
               "MP": cur.get("mp", "")}
        for s in SPOTS:
            rec[f"{s}%"] = cur.get(s)
        rows.append(rec)
    rows.sort(key=lambda x: (x["Agrees"] == "yes", -_num(x["MP"])))
    n_no = sum(1 for x in rows if x["Agrees"] == "no")
    print(f"  league audit: {len(rows)} rostered players with 2025-26 minutes, "
          f"{n_no} disagree with the master roster")
    return rows


def collect(career: bool = True) -> list:
    """Season percentages first (one page for the league), then career per
    player. Pass career=False for the quick pass while the slow one runs."""
    print(f"{SEASON_LABEL} play-by-play, league-wide")
    cur_all = season_pcts(SEASON_END)
    print(f"  {len(cur_all)} players")
    AUDIT[:] = audit(cur_all)
    ids = player_ids(SEASON_END) if career else {}
    out = []
    for name, assigned in PLAYERS.items():
        cur = cur_all.get(normalize(name), {})
        rec = {"Player": name, "Team 25-26": cur.get("team", ""), "Assigned": assigned,
               f"{SEASON_LABEL} top 2": top_two(cur), f"{SEASON_LABEL} MP": cur.get("mp", "")}
        for s in SPOTS:
            rec[f"{SEASON_LABEL} {s}%"] = cur.get(s)
        rec[f"Matches {SEASON_LABEL}"] = ("yes" if rec[f"{SEASON_LABEL} top 2"] == assigned
                                          else "no")
        car, pid = {}, ids.get(normalize(name), "")
        if career and pid:
            rows, keys = pbp_rows(pid)
            career_row = next((r for r in rows
                               if re.match(r"^\d+\s+Yrs?$", _label(r).strip())), None)
            car = pct(career_row, keys) if career_row else {}
            rec["Career MP"] = (career_row or {}).get("mp", "")
        rec["BBRef"] = pid
        rec["Career top 2"] = top_two(car)
        for s in SPOTS:
            rec[f"Career {s}%"] = car.get(s)
        rec["Matches career"] = "yes" if rec["Career top 2"] == assigned else "no"
        out.append(rec)
        print(f"  {name:20} assigned {assigned:6} | career {rec['Career top 2'] or '-':6}"
              f" | {SEASON_LABEL} {rec[f'{SEASON_LABEL} top 2'] or '-'}")
    return out


def write_xlsx(rows: list) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Position estimates"
    cols = ["Player", "Team 25-26", "Assigned", "Career top 2", f"{SEASON_LABEL} top 2",
            "Matches career", f"Matches {SEASON_LABEL}",
            *[f"Career {s}%" for s in SPOTS], "Career MP",
            *[f"{SEASON_LABEL} {s}%" for s in SPOTS], f"{SEASON_LABEL} MP", "BBRef"]
    F = "Arial"
    head = Font(name=F, size=10, bold=True, color="FFFFFF")
    base = Font(name=F, size=10)
    bold = Font(name=F, size=10, bold=True)
    fill_h = PatternFill("solid", fgColor="1F3864")
    fill_car = PatternFill("solid", fgColor="EDF2FA")
    fill_yes = PatternFill("solid", fgColor="D9EAD3")
    fill_no = PatternFill("solid", fgColor="FCE4E4")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(cols)
    for i, c in enumerate(ws[1], 1):
        c.font, c.fill, c.border = head, fill_h, border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font, c.border = base, border
            name = cols[c.column - 1]
            if name.startswith("Career") and name.endswith("%"):
                c.fill = fill_car
            if name.endswith("%"):
                c.number_format = "0"
                c.alignment = Alignment(horizontal="center")
            if name in ("Assigned", "Career top 2", f"{SEASON_LABEL} top 2"):
                c.font = bold
                c.alignment = Alignment(horizontal="center")
            if name.startswith("Matches"):
                c.alignment = Alignment(horizontal="center")
                if c.value == "yes":
                    c.fill = fill_yes
                elif c.value == "no":
                    c.fill = fill_no
    widths = {"Player": 20, "Team 25-26": 11, "Assigned": 10, "Career top 2": 12,
              f"{SEASON_LABEL} top 2": 13, "Matches career": 14,
              f"Matches {SEASON_LABEL}": 15, "Career MP": 10,
              f"{SEASON_LABEL} MP": 12, "BBRef": 12}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths.get(c, 9)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    # Sheet 2: the same check across the whole league. The season page is one
    # request for every player, so auditing all 439 rostered players costs
    # nothing extra, and it is the list to work from for the rest of the pass.
    if AUDIT:
        wa = wb.create_sheet("Every rostered player")
        acols = ["Player", "Team", "Master pos", f"{SEASON_LABEL} top 2", "Agrees",
                 *[f"{s}%" for s in SPOTS], "MP"]
        wa.append(acols)
        for i, c in enumerate(wa[1], 1):
            c.font, c.fill, c.border = head, fill_h, border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in AUDIT:
            wa.append([r.get(c, "") for c in acols])
        for row in wa.iter_rows(min_row=2):
            for c in row:
                c.font, c.border = base, border
                nm = acols[c.column - 1]
                if nm.endswith("%"):
                    c.number_format = "0"
                    c.alignment = Alignment(horizontal="center")
                if nm in ("Master pos", f"{SEASON_LABEL} top 2"):
                    c.font = bold
                    c.alignment = Alignment(horizontal="center")
                if nm == "Agrees":
                    c.alignment = Alignment(horizontal="center")
                    if c.value == "yes":
                        c.fill = fill_yes
                    elif c.value == "no":
                        c.fill = fill_no
        for i, c in enumerate(acols, 1):
            wa.column_dimensions[wa.cell(row=1, column=i).column_letter].width = (
                {"Player": 22, "Team": 8, "Master pos": 11,
                 f"{SEASON_LABEL} top 2": 13, "Agrees": 9, "MP": 8}.get(c, 7))
        wa.freeze_panes = "B2"
        wa.auto_filter.ref = wa.dimensions

    ws2 = wb.create_sheet("How to read this")
    for line in [
        ["Source", "Basketball Reference, play-by-play table, Position Estimate columns"],
        ["What the percentages are",
         "The share of a player's minutes spent at each spot, as BBRef estimates it "
         "from play-by-play data. Rows do not always total exactly 100."],
        ["Career", "BBRef's own career row, minutes-weighted across every season "
                   "since 1996-97 (play-by-play data does not exist before that)."],
        [SEASON_LABEL, "That season's row only."],
        ["Top 2", "The two spots with the most minutes, primary first. A single "
                  "position is shown when the second spot is under 10 percent."],
        ["Assigned", "The position currently on data/master_roster.csv."],
        ["Matches", "Whether the assigned pair equals the top-2 pair, order included."],
    ]:
        ws2.append(line)
    for row in ws2.iter_rows():
        for c in row:
            c.font = bold if c.column == 1 else base
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 95

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    data = collect(career="--season-only" not in sys.argv)
    (CACHE / "_last_run.json").write_text(json.dumps(data, indent=1))
    write_xlsx(data)
