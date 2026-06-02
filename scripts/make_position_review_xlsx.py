"""Build player_positions_review.xlsx: every rostered player, grouped by team,
then by position group (Guards -> Forwards -> Centers), Barrett-descending within
each group. Separate Primary/Secondary columns (with dropdowns) for the user to
correct, plus a Source column (override / 2K / fallback-guess). Self-locating.

Usage:  python -u scripts/make_position_review_xlsx.py
"""
import os, sys, warnings
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT); os.chdir(ROOT)
warnings.filterwarnings("ignore")
import pandas as pd
from utils import build_ranked_projected, normalize, SEASONS
import team_suitors as ts

SEASON = max(SEASONS, key=lambda s: int(s[:4]))
print("season:", SEASON, flush=True)
ranked = build_ranked_projected(SEASON)
posmap = ts.load_player_positions()

# source sets: which file a player's position came from
def _names(path):
    try:
        df = pd.read_csv(path, comment="#", skip_blank_lines=True)
        return {normalize(n) for n in df["name"]}
    except Exception:
        return set()
ovr_names, twok_names = _names(ts.POS_OVERRIDE_PATH), _names(ts.POS_2K_PATH)

pick = lambda *c: next((x for x in c if x in ranked.columns), None)
tcol = pick("Team", "TEAM_ABBREVIATION", "team")
pcol = pick("Player", "PLAYER_NAME", "name")
bcol = pick("barrett_score", "barrett", "Barrett")
poscol = pick("position_detailed", "position", "POSITION", "Pos")

_GRP = {"PG": ("Guard", 0), "SG": ("Guard", 0), "SF": ("Forward", 1),
        "PF": ("Forward", 1), "C": ("Center", 2)}
rows = []
for _, r in ranked.iterrows():
    name = str(r[pcol]); team = str(r[tcol]) if tcol and pd.notna(r[tcol]) else "FA"
    if not team or team == "nan":
        team = "FA"
    bbref = str(r[poscol]) if poscol and pd.notna(r[poscol]) else ""
    resolved = ts.resolve_position(name, bbref, posmap)            # primary-first
    parts = [p for p in str(resolved).replace("|", "/").split("/") if p.strip()]
    primary = ts._norm5(parts[0]) if parts else "SF"
    secondary = ts._norm5(parts[1]) if len(parts) > 1 else ""
    grp, grank = _GRP.get(primary, ("Forward", 1))
    nn = normalize(name)
    source = "override" if nn in ovr_names else ("2K" if nn in twok_names else "fallback")
    try:
        bar = round(float(r[bcol]), 1)
    except Exception:
        bar = None
    rows.append({"Team": team, "Player": name, "Group": grp, "Primary": primary,
                 "Secondary": secondary, "Barrett": bar, "Source": source,
                 "_grank": grank, "_bar": bar if bar is not None else -999})

df = pd.DataFrame(rows).sort_values(
    ["Team", "_grank", "_bar"], ascending=[True, True, False]).drop(columns=["_grank", "_bar"])
print(f"{len(df)} players; sources: " + ", ".join(f"{k}={v}" for k, v in df['Source'].value_counts().items()), flush=True)

OUT = ROOT + "/player_positions_review.xlsx"
df.to_excel(OUT, index=False, sheet_name="positions")

# formatting + dropdowns
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
wb = load_workbook(OUT); ws = wb["positions"]
n = len(df)
hdr_fill = PatternFill("solid", fgColor="1f6fb4"); flag = PatternFill("solid", fgColor="fff3cd")
for c in range(1, 8):
    cell = ws.cell(1, c); cell.font = Font(bold=True, color="ffffff"); cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center")
ws.freeze_panes = "A2"
widths = {"A": 7, "B": 26, "C": 10, "D": 9, "E": 11, "F": 9, "G": 10}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
dv_p = DataValidation(type="list", formula1='"PG,SG,SF,PF,C"', allow_blank=False)
dv_s = DataValidation(type="list", formula1='"PG,SG,SF,PF,C"', allow_blank=True)
ws.add_data_validation(dv_p); ws.add_data_validation(dv_s)
dv_p.add(f"D2:D{n+1}"); dv_s.add(f"E2:E{n+1}")
for i in range(2, n + 2):                                   # highlight fallback rows for review
    if ws.cell(i, 7).value == "fallback":
        ws.cell(i, 7).fill = flag
wb.save(OUT)
print(f"saved -> {OUT}", flush=True)
