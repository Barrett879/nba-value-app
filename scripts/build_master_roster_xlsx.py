"""Build the Master Roster workbook from data/master_roster.csv.

The CSV is the maintained source of truth (same doctrine as real_signings:
hand-edit as trades/waivers happen, re-run this). Output workbook:
  Sheet 1  Master List   - every player on every team, filterable
  Sheet 2  Team Summary  - per-team counts + payroll
  Sheet 3  About         - as-of date, sources, how to maintain

Run: python3 scripts/build_master_roster_xlsx.py [out.xlsx]
Default output: exports/HoopsValue_Master_Roster.xlsx (NOT committed; the CSV is).
"""
import csv
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "master_roster.csv"
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "exports" / "HoopsValue_Master_Roster.xlsx"

FONT = "Arial"
base = Font(name=FONT, size=10)
bold = Font(name=FONT, size=10, bold=True)
hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F3A5F")
band = PatternFill("solid", fgColor="EEF2F7")
team_fill = PatternFill("solid", fgColor="DCE6F2")
pend_font = Font(name=FONT, size=10, italic=True, color="8A6D1A")
two_font = Font(name=FONT, size=10, color="4F6B8A")
fa_font = Font(name=FONT, size=10, italic=True, color="6B7280")


def main() -> None:
    rows = []
    asof = ""
    with SRC.open() as f:
        for r in csv.DictReader(x for x in f if not x.lstrip().startswith("#")):
            if r.get("asof"):
                asof = r["asof"] or asof
            rows.append(r)
    rows.sort(key=lambda r: (r["team"], {"standard": 0, "two_way": 1, "pending": 2, "free_agent": 3}[r["kind"]],
                             -float(r["salary_M"] or 0)))

    wb = Workbook()

    # ── Master List ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Master List"
    heads = ["Team", "Player", "Pos", "2026-27 Salary ($M)", "Contract", "Notes"]
    ws.append(heads)
    for c in range(1, len(heads) + 1):
        cell = ws.cell(1, c)
        cell.font = hdr_font
        cell.fill = hdr_fill
    kind_label = {"standard": "Standard", "two_way": "Two-Way", "pending": "PENDING",
                  "free_agent": "Free Agent"}
    prev_team = None
    for i, r in enumerate(rows, start=2):
        ws.append([r["team"], r["player"], r["pos"],
                   float(r["salary_M"] or 0), kind_label[r["kind"]], r["notes"]])
        first_of_team = r["team"] != prev_team
        prev_team = r["team"]
        for c in range(1, len(heads) + 1):
            cell = ws.cell(i, c)
            cell.font = (pend_font if r["kind"] == "pending"
                         else two_font if r["kind"] == "two_way"
                         else fa_font if r["kind"] == "free_agent" else base)
            if first_of_team:
                cell.fill = team_fill
            elif i % 2 == 0:
                cell.fill = band
        if first_of_team:
            ws.cell(i, 1).font = bold
        ws.cell(i, 4).number_format = "0.00"
    for j, w in enumerate([7, 24, 8, 17, 11, 60], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

    # ── Team Summary ─────────────────────────────────────────────────────────
    ts = wb.create_sheet("Team Summary")
    ts.append(["Team", "Standard", "Two-Way", "Pending", "Free Agents", "Payroll ($M)", "Top salaries"])
    for c in range(1, 8):
        ts.cell(1, c).font = hdr_font
        ts.cell(1, c).fill = hdr_fill
    per = defaultdict(list)
    for r in rows:
        per[r["team"]].append(r)
    for i, t in enumerate(sorted(per), start=2):
        std = [r for r in per[t] if r["kind"] == "standard"]
        tw = [r for r in per[t] if r["kind"] == "two_way"]
        pend = [r for r in per[t] if r["kind"] == "pending"]
        fa = [r for r in per[t] if r["kind"] == "free_agent"]
        pay = sum(float(r["salary_M"] or 0) for r in std)
        top = ", ".join(f'{r["player"]} {float(r["salary_M"]):.1f}'
                        for r in sorted(std, key=lambda x: -float(x["salary_M"] or 0))[:3])
        ts.append([t, len(std), len(tw), len(pend), len(fa), round(pay, 1), top])
        for c in range(1, 8):
            ts.cell(i, c).font = base
            if i % 2 == 0:
                ts.cell(i, c).fill = band
        ts.cell(i, 1).font = bold
        if len(std) > 15:
            ts.cell(i, 2).font = Font(name=FONT, size=10, bold=True, color="B00020")
    for j, w in enumerate([7, 10, 9, 9, 12, 13, 58], 1):
        ts.column_dimensions[get_column_letter(j)].width = w
    ts.freeze_panes = "A2"

    # ── About ────────────────────────────────────────────────────────────────
    ab = wb.create_sheet("About")
    lines = [
        ("HoopsValue Master Roster", bold),
        (f"As of: {asof}", base),
        ("", base),
        ("Every player under contract for 2026-27, all 30 teams.", base),
        ("Compiled by six division research agents from Spotrac (canonical),", base),
        ("cross-checked against ESPN and RealGM; conflicts resolved with", base),
        ("transaction-level checks. PENDING rows are agreed-but-not-official", base),
        ("deals and do not count toward roster limits. FREE AGENT rows are the", base),
        ("verified unsigned 2026 FA registry, listed under the team holding", base),
        ("their rights; they feed the Trade Machine sign-and-trade lists.", base),
        ("", base),
        ("To maintain: edit data/master_roster.csv as trades and signings", base),
        ("happen, then re-run scripts/build_master_roster_xlsx.py.", base),
        ("This file is the roster source of truth for the site pipeline.", base),
    ]
    for i, (txt, f) in enumerate(lines, start=1):
        ab.cell(i, 1, txt).font = f
    ab.column_dimensions["A"].width = 70

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    n_std = sum(1 for r in rows if r["kind"] == "standard")
    n_tw = sum(1 for r in rows if r["kind"] == "two_way")
    n_p = sum(1 for r in rows if r["kind"] == "pending")
    n_fa = sum(1 for r in rows if r["kind"] == "free_agent")
    print(f"wrote {OUT} ({len(rows)} rows: {n_std} standard, {n_tw} two-way, "
          f"{n_p} pending, {n_fa} free agents)")


if __name__ == "__main__":
    main()
