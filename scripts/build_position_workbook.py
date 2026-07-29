#!/usr/bin/env python3
"""Every player on the 2026-27 roster, checked against what he actually plays.

Basketball Reference estimates the share of a player's minutes spent at each
of the five spots. This pulls that for the whole league, career and recent,
and sets it against the position on data/master_roster.csv.

Career comes from one league page per season since 2003-04 rather than one page
per player: BBRef's own career row is a minutes-weighted average of the seasons,
so aggregating the season pages reproduces it for 23 requests instead of 440.

"Recent" is the most recent season the player cleared 500 minutes, so a
5-game season does not get to decide what position he plays.

    python scripts/build_position_workbook.py
"""
import csv
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "scripts"))

from utils import normalize  # noqa: E402
import fetch_bref_position_pct as B  # noqa: E402

OUT = ROOT / "exports" / "HoopsValue_Position_Estimates.xlsx"
FLOOR = 500
SPOTS = B.SPOTS
# the positions set by hand this session, kept on their own sheet
ASSIGNED = B.PLAYERS


def verdict(master: str, top: str) -> str:
    """How far apart the hand-set position and the measured one are. A flat
    yes/no is not actionable when the two come from different taxonomies: 2K
    labels what a player is, BBRef measures where he stood."""
    if not top or not master:
        return ""
    if top == master:
        return "match"
    m, t = master.split("/"), top.split("/")
    if len(m) == 2 and len(t) == 2 and set(m) == set(t):
        return "order flipped"
    if m[0] == t[0]:
        return "primary ok"
    if len(t) == 1 and t[0] in m:
        return "primary ok" if m[0] == t[0] else "second only"
    if len(m) == 2 and t[0] == m[1]:
        return "second only"
    return "different"


def likely_errors(rows: list, dominant: int = 55) -> list:
    """Rows where the master position is not a judgment call but wrong.

    The bar: the player spent a dominant share of his career at a spot that is
    not even listed on the master roster. Terry Rozier at 65 percent PG is not
    a SF/PF by anyone's taxonomy. Judgment calls (order, second spot) are left
    out; this is the list you can act on without arguing.
    """
    out = []
    for r in rows:
        top, master = r["Career top 2"], r["Master pos"]
        if not top or not master:
            continue
        primary = top.split("/")[0]
        share = r.get(f"Career {primary}%") or 0
        if primary not in master.split("/") and share >= dominant:
            out.append(r)
    out.sort(key=lambda x: -(x["Career MP"] or 0))
    return out


def roster() -> list:
    lines = [l for l in (ROOT / "data" / "master_roster.csv").read_text().splitlines()
             if not l.lstrip().startswith("#")]
    return [r for r in csv.DictReader(lines) if (r.get("player") or "").strip()]


def build() -> list:
    career, hist = B.career_pcts(B.FIRST_SEASON_END, B.SEASON_END)
    print(f"career: {len(career)} players across "
          f"{B.FIRST_SEASON_END - 1}-{B.SEASON_END}")
    rows = []
    for r in roster():
        name = r["player"].strip()
        key = normalize(name)
        car = career.get(key) or {}
        end, rec = B.recent_qualifying(hist.get(key), FLOOR)
        rec = rec or {}
        master = (r.get("pos") or "").strip()
        car_top, rec_top = B.top_two(car), B.top_two(rec)
        season = f"{end - 1}-{str(end)[2:]}" if end else ""
        row = {
            "Player": name, "Team": (r.get("team") or "").strip(),
            "Kind": (r.get("kind") or "").strip().replace("_", "-"),
            "Master pos": master,
            "Career top 2": car_top, "Recent top 2": rec_top,
            "Recent season": season,
            "Career verdict": verdict(master, car_top),
            "Recent verdict": verdict(master, rec_top),
        }
        for s in SPOTS:
            row[f"Career {s}%"] = car.get(s)
        row["Career MP"] = car.get("mp")
        for s in SPOTS:
            row[f"Recent {s}%"] = rec.get(s)
        row["Recent MP"] = B._num(rec.get("mp")) or None
        row["Hand set"] = "yes" if name in ASSIGNED else ""
        rows.append(row)
    order = {"different": 0, "second only": 1, "primary ok": 2, "order flipped": 3,
             "match": 4, "": 5}
    rows.sort(key=lambda x: (order.get(x["Career verdict"], 5), -(x["Career MP"] or 0)))
    return rows


def write(rows: list) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    cols = ["Player", "Team", "Kind", "Master pos", "Career top 2", "Recent top 2",
            "Recent season", "Career verdict", "Recent verdict",
            *[f"Career {s}%" for s in SPOTS], "Career MP",
            *[f"Recent {s}%" for s in SPOTS], "Recent MP", "Hand set"]
    F = "Arial"
    head = Font(name=F, size=10, bold=True, color="FFFFFF")
    base, bold = Font(name=F, size=10), Font(name=F, size=10, bold=True)
    fill_h = PatternFill("solid", fgColor="1F3864")
    fill_car = PatternFill("solid", fgColor="EDF2FA")
    fill_yes = PatternFill("solid", fgColor="D9EAD3")
    fill_no = PatternFill("solid", fgColor="FCE4E4")
    fill_warn = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = {"Player": 22, "Team": 7, "Kind": 10, "Master pos": 11, "Career top 2": 12,
              "Recent top 2": 12, "Recent season": 13, "Career verdict": 14,
              "Recent verdict": 14, "Career MP": 10, "Recent MP": 10, "Hand set": 9}

    def sheet(ws, data):
        ws.append(cols)
        for c in ws[1]:
            c.font, c.fill, c.border = head, fill_h, border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in data:
            ws.append([r.get(c) for c in cols])
        for row in ws.iter_rows(min_row=2):
            for c in row:
                name = cols[c.column - 1]
                c.font, c.border = base, border
                if name.startswith("Career") and name.endswith("%"):
                    c.fill = fill_car
                if name.endswith("%"):
                    c.number_format = "0"
                    c.alignment = Alignment(horizontal="center")
                if name in ("Master pos", "Career top 2", "Recent top 2"):
                    c.font = bold
                    c.alignment = Alignment(horizontal="center")
                if name.endswith("verdict") or name in ("Kind", "Team", "Hand set",
                                                        "Recent season"):
                    c.alignment = Alignment(horizontal="center")
                if name.endswith("verdict"):
                    if c.value == "match":
                        c.fill = fill_yes
                    elif c.value in ("different", "second only"):
                        c.fill = fill_no
                    elif c.value in ("order flipped", "primary ok"):
                        c.fill = fill_warn
        for i, c in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
                widths.get(c, 8)
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

    wb = Workbook()
    sheet(wb.active, rows)
    wb.active.title = "Every rostered player"
    sheet(wb.create_sheet("Likely errors"), likely_errors(rows))
    sheet(wb.create_sheet("Hand set this session"),
          [r for r in rows if r["Hand set"] == "yes"])

    ws = wb.create_sheet("How to read this")
    for line in [
        ["Source", "Basketball Reference, Position Estimate columns on the "
                   "play-by-play tables"],
        ["What the percentages are",
         "The share of a player's minutes spent at each spot, as BBRef estimates "
         "it from play-by-play data. Rows do not always total exactly 100."],
        ["Career", f"Minutes-weighted across every season from "
                   f"{B.FIRST_SEASON_END - 1}-{str(B.FIRST_SEASON_END)[2:]} on. "
                   f"Aggregated from the league page for each season, which is "
                   f"how BBRef builds its own career row."],
        ["Recent", f"The most recent season the player cleared {FLOOR} minutes, "
                   f"named in the Recent season column. A player who has never "
                   f"cleared it shows his latest season instead."],
        ["Top 2", "The two spots with the most minutes, primary first. A single "
                  "position is shown when the second spot is under 10 percent."],
        ["Master pos", "The position currently on data/master_roster.csv."],
        ["Verdict", "match = same pair, same order. order flipped = same two "
                    "spots, primary and secondary swapped. primary ok = the "
                    "primary agrees, the second spot does not. second only = "
                    "the measured primary is the master's SECOND position. "
                    "different = they do not overlap at the primary. Blank means "
                    "BBRef has no play-by-play data for him, which is every "
                    "player who has not played an NBA game."],
        ["Sort", "Disagreements first, then by career minutes, so the players "
                 "worth looking at are at the top."],
    ]:
        ws.append(line)
    for row in ws.iter_rows():
        for c in row:
            c.font = bold if c.column == 1 else base
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 95

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    data = build()
    n = len(data)
    have = [r for r in data if r["Career top 2"]]
    from collections import Counter
    cc, cr = Counter(r["Career verdict"] for r in have), Counter(r["Recent verdict"] for r in have)
    print(f"\n{n} players on the master roster, {len(have)} with NBA play-by-play data")
    for lbl, c in (("career", cc), ("recent", cr)):
        parts = ", ".join(f"{k or 'no data'} {v}" for k, v in c.most_common())
        print(f"  {lbl}: {parts}")
    le = likely_errors(data)
    print(f"\n  likely errors (dominant career spot not listed at all): {len(le)}")
    for r in le[:25]:
        p = r["Career top 2"].split("/")[0]
        print(f"    {r['Player']:24} {r['Team']:4} master {r['Master pos']:8}"
              f" plays {p} {int(r[f'Career {p}%'])}% of {r['Career MP']} min")
    write(data)
