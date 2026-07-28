"""Build the Trade Exceptions workbook from data/trade_exceptions_2026_27.csv.

The CSV is the maintained source of truth (edit as TPEs are created, used, or
expire; re-run this). Output workbook:
  Sheet 1  Trade Exceptions - every outstanding TPE, largest first per team
  Sheet 2  About            - as-of date, sources, how to maintain

Run: python3 scripts/build_trade_exceptions_xlsx.py [out.xlsx]
Default output: exports/HoopsValue_Trade_Exceptions.xlsx (NOT committed; the CSV is).
"""
import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "trade_exceptions_2026_27.csv"
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "exports" / "HoopsValue_Trade_Exceptions.xlsx"

FONT = "Arial"
base = Font(name=FONT, size=10)
bold = Font(name=FONT, size=10, bold=True)
hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F3A5F")
band = PatternFill("solid", fgColor="EEF2F7")
team_fill = PatternFill("solid", fgColor="DCE6F2")
big_font = Font(name=FONT, size=10, bold=True, color="1A6B3C")


def main() -> None:
    rows, asof = [], ""
    with SRC.open() as f:
        for r in csv.DictReader(x for x in f if not x.lstrip().startswith("#")):
            if r.get("asof"):
                asof = r["asof"] or asof
            if r.get("team"):
                rows.append(r)
    rows.sort(key=lambda r: (r["team"], -float(r["amount_M"] or 0)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Exceptions"
    heads = ["Team", "Amount ($M)", "From player", "Created", "Expires", "Notes"]
    ws.append(heads)
    for c in range(1, len(heads) + 1):
        ws.cell(1, c).font = hdr_font
        ws.cell(1, c).fill = hdr_fill
    prev = None
    for i, r in enumerate(rows, start=2):
        amt = float(r["amount_M"] or 0)
        ws.append([r["team"], amt, r.get("player", ""), r.get("created", ""),
                   r.get("expires", ""), r.get("notes", "")])
        first = r["team"] != prev
        prev = r["team"]
        for c in range(1, len(heads) + 1):
            cell = ws.cell(i, c)
            cell.font = base
            if first:
                cell.fill = team_fill
            elif i % 2 == 0:
                cell.fill = band
        if first:
            ws.cell(i, 1).font = bold
        if amt >= 10:
            ws.cell(i, 2).font = big_font
        ws.cell(i, 2).number_format = "0.00"
    for j, w in enumerate([7, 13, 24, 12, 12, 64], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

    ab = wb.create_sheet("About")
    lines = [
        ("HoopsValue Trade Exceptions Master List", bold),
        (f"As of: {asof}", base),
        ("", base),
        ("Every outstanding traded-player exception (TPE) in the league,", base),
        ("verified against the RealGM trade-exception tracker, Spotrac team", base),
        ("cap pages, and Hoops Rumors reporting. TPEs last one year from", base),
        ("creation and shrink as they absorb salary; amounts are current", base),
        ("remaining balances.", base),
        ("", base),
        ("Cap notes: using a TPE created in a PRIOR season hard-caps the team", base),
        ("at the first apron. Teams above the first apron cannot use a", base),
        ("prior-season TPE at all.", base),
        ("", base),
        ("To maintain: edit data/trade_exceptions_2026_27.csv as TPEs are", base),
        ("created, used, or expire, then re-run this script. The list also", base),
        ("shows on Trade Machine team panels.", base),
    ]
    for i, (txt, f) in enumerate(lines, start=1):
        ab.cell(i, 1, txt).font = f
    ab.column_dimensions["A"].width = 70

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    total = sum(float(r["amount_M"] or 0) for r in rows)
    print(f"wrote {OUT} ({len(rows)} TPEs across "
          f"{len(set(r['team'] for r in rows))} teams, ${total:.1f}M total)")


if __name__ == "__main__":
    main()
