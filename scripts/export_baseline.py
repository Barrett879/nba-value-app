#!/usr/bin/env python3
"""Baseline roster builder + free-agent assignment grid.

Splits the resolved projection (HoopsValue_FreeAgency_2026.xlsx) into:
  - ROSTERS: kept options + guaranteed deals + draft picks (the option decisions the
    model / the user / the articles already made are respected - declined and opted-out
    players are NOT here), with NO re-signs or signings.
  - ASSIGN FREE AGENTS: every re-sign / new signing / free agent, with empty "Sign With"
    and "Salary" columns to fill in by hand (the model's pick is shown as a hint).
  - TEAM ROOM: each team's open spots and money available after the baseline.

Reads the resolved projection so all option calls flow through automatically.
Usage:  python scripts/export_baseline.py   (run export_fa_spreadsheet.py first)
"""
import csv
import datetime
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
from utils import normalize  # noqa: E402

# Official signings (data/manual_signings.csv) lock onto rosters, not the assign grid.
_ms = ROOT / "data" / "manual_signings.csv"
ms_names = set()
if _ms.exists():
    for r in csv.DictReader(l for l in _ms.read_text().splitlines() if l.strip() and not l.lstrip().startswith("#")):
        if (r.get("player") or "").strip():
            ms_names.add(normalize(r["player"].strip()))

SRC = Path.home() / "Downloads" / "HoopsValue_FreeAgency_2026.xlsx"
wb_in = load_workbook(SRC)
d = wb_in["Roster Detail"]
dh = {(d.cell(1, j).value or "").replace("\n", " "): j for j in range(1, d.max_column + 1)}
TI, PI, POSI, TYI, STI, PREV, BAR, COST = (dh["Team"], dh["Player"], dh["Position"], dh["Type"],
                                           dh["Status"], dh["Previous Team"], dh["Barrett Score"], dh["Cost / Salary ($M)"])
fa = wb_in["2026 Free Agents"]
fh = {(fa.cell(4, j).value or "").replace("\n", " "): j for j in range(1, fa.max_column + 1)}

CAP, TAX, AP1, AP2 = 165, 200, 209, 222

# Split the detail: kept-on-roster (options + guaranteed + picks) vs FA-market (re-sign / signing).
roster_rows, grid = [], []                                                       # grid = [player,pos,prev,bar,value,suggestion]
for r in range(2, d.max_row + 1):
    typ = d.cell(r, TYI).value
    official = normalize((d.cell(r, PI).value or "").strip()) in ms_names
    if typ in ("Under contract", "Draft pick") or (typ in ("Re-sign", "New signing") and official):
        label = typ if official else d.cell(r, STI).value                        # show the move, not the pre-signing "UFA"
        row = [d.cell(r, TI).value, d.cell(r, PI).value, d.cell(r, POSI).value, label, d.cell(r, BAR).value, d.cell(r, COST).value]
        roster_rows.append(row + [typ])                                          # official signings lock to the roster
    elif typ in ("Re-sign", "New signing"):
        grid.append([d.cell(r, PI).value, d.cell(r, POSI).value, d.cell(r, PREV).value,
                     d.cell(r, BAR).value, d.cell(r, COST).value, d.cell(r, TI).value])
for r in range(5, fa.max_row + 1):
    if not fa.cell(r, fh["Player"]).value:
        continue
    grid.append([fa.cell(r, fh["Player"]).value, fa.cell(r, fh["Position"]).value, fa.cell(r, fh["Current Team"]).value,
                 fa.cell(r, fh["Barrett Score"]).value, fa.cell(r, fh["Est. Value ($M)"]).value, fa.cell(r, fh["Projected Landing"]).value])
grid.sort(key=lambda g: -(g[4] or g[3] or 0))                                    # by value, then barrett

# Per-team baseline aggregates (from roster rows only)
team_order = sorted({r[0] for r in roster_rows})
agg = {}
for tm in team_order:
    rs = [r for r in roster_rows if r[0] == tm]
    std = sum(1 for r in rs if r[2] != "2nd round")
    total = round(sum(r[5] or 0 for r in rs), 1)
    agg[tm] = {"size": std, "open": max(0, 15 - std), "total": total}


def _power(c):
    cap_room = round(CAP - c, 1)
    if cap_room >= 8:    return max(cap_room, 0.0), "Cap room"
    if c >= AP2:         return 0.0, "Minimums only"
    if c >= AP1:         return 5.7, "Taxpayer MLE"
    return 14.1, "Full MLE"


FONT = "Arial"
base, boldf = Font(name=FONT, size=10), Font(name=FONT, size=10, bold=True)
hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F3A5F")
edit_fill = PatternFill("solid", fgColor="FFF6D6")
edit_hdr = Font(name=FONT, size=10, bold=True, color="1F3A5F")
band = PatternFill("solid", fgColor="EEF2F7")
MONEY, SCORE = '$#,##0.0;($#,##0.0);"-"', '0.0'
rt, ct = Alignment(horizontal="right"), Alignment(horizontal="center")
wrap_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
team_top = Border(top=Side(style="medium", color="9AA7B5"))
wb = Workbook()

# Sheet 1: Rosters
s = wb.active; s.title = "Rosters"
s["A1"] = "Baseline rosters - kept options + guaranteed deals + draft picks + official signings"
s["A1"].font = Font(name=FONT, size=13, bold=True, color="1F3A5F")
s["A2"] = "Option decisions are as projected (declined / opted-out players are in the free-agent list, not here). Reported signings are locked in; assign the rest on the next sheet."
s["A2"].font = Font(name=FONT, size=9, italic=True, color="6B7682")
cols = ["Team", "Player", "Position", "Type", "Barrett", "Salary ($M)"]
for j, c in enumerate(cols, 1):
    cell = s.cell(4, j, c); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = wrap_c
for i, r in enumerate(roster_rows):
    rr = 5 + i
    for j, v in enumerate(r[:6], 1):
        cell = s.cell(rr, j, v); cell.font = base
        if j == 5:
            cell.number_format = SCORE; cell.alignment = ct
        elif j == 6:
            cell.number_format = MONEY; cell.alignment = rt
        elif j in (1, 4):
            cell.alignment = ct
    if i > 0 and r[0] != roster_rows[i - 1][0]:
        for j in range(1, 7):
            s.cell(rr, j).border = team_top
for j, w in enumerate([7, 24, 11, 16, 8, 12], 1):
    s.column_dimensions[get_column_letter(j)].width = w
s.freeze_panes = "C5"

# Sheet 2: Assign Free Agents
a = wb.create_sheet("Assign Free Agents")
a["A1"] = "Assign Free Agents - fill in the two yellow columns"
a["A1"].font = Font(name=FONT, size=13, bold=True, color="1F3A5F")
a["A2"] = ("Type a team's 3-letter abbreviation in 'Sign With' and a salary in $M for anyone you want signed; leave blank "
           "to keep him a free agent. 'Model pick' is just the projection's guess for reference. Sorted best-first.")
a["A2"].font = Font(name=FONT, size=9, italic=True, color="6B7682")
gcols = ["Player", "Pos", "Prev Team", "Barrett", "Est. Value\n($M)", "Model pick", "Sign With\n(team)", "Salary\n($M)"]
for j, c in enumerate(gcols, 1):
    cell = a.cell(4, j, c); cell.alignment = wrap_c
    if j >= 7:
        cell.fill = edit_fill; cell.font = edit_hdr
    else:
        cell.fill = hdr_fill; cell.font = hdr_font
for i, g in enumerate(grid):
    rr = 5 + i
    vals = list(g) + [None, None]                                                # + Sign With, Salary
    for j, v in enumerate(vals, 1):
        cell = a.cell(rr, j, v); cell.font = base
        if j == 4:
            cell.number_format = SCORE; cell.alignment = ct
        elif j == 5:
            cell.number_format = MONEY; cell.alignment = rt
        elif j in (2, 3, 6, 7):
            cell.alignment = ct
        if j >= 7:
            cell.fill = edit_fill
    if i % 2 == 1:
        for j in range(1, 7):
            a.cell(rr, j).fill = band
for j, w in enumerate([24, 9, 10, 8, 11, 11, 12, 10], 1):
    a.column_dimensions[get_column_letter(j)].width = w
a.freeze_panes = "B5"

# Sheet 3: Team Room
tr = wb.create_sheet("Team Room")
tr["A1"] = "Team Room - open spots and money available after the baseline"
tr["A1"].font = Font(name=FONT, size=13, bold=True, color="1F3A5F")
tr["A2"] = "Spending Power = the realistic tool to add an outside FA (cap room, or the mid-level by apron tier)."
tr["A2"].font = Font(name=FONT, size=9, italic=True, color="6B7682")
tcols = ["Abbr", "Roster", "Open\nSpots", "Committed\n($M)", "Cap Space\n($M)", "Spending\nPower ($M)",
         "Tier", "Room to\nTax ($M)", "Room to\n2nd Apron ($M)"]
for j, c in enumerate(tcols, 1):
    cell = tr.cell(4, j, c); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = wrap_c
for i, tm in enumerate(team_order):
    g = agg[tm]; power, tier = _power(g["total"]); rr = 5 + i
    vals = [tm, g["size"], g["open"], g["total"], round(CAP - g["total"], 1), power, tier,
            round(TAX - g["total"], 1), round(AP2 - g["total"], 1)]
    for j, v in enumerate(vals, 1):
        cell = tr.cell(rr, j, v); cell.font = base
        if j in (4, 5, 6, 8, 9):
            cell.number_format = MONEY; cell.alignment = rt
        elif j in (1, 2, 3, 7):
            cell.alignment = ct
    if i % 2 == 1:
        for j in range(1, 10):
            tr.cell(rr, j).fill = band
for j, w in enumerate([7, 7, 7, 12, 11, 11, 14, 11, 14], 1):
    tr.column_dimensions[get_column_letter(j)].width = w
tr.freeze_panes = "B5"

stamp = datetime.datetime.now().strftime("%b%d_%H%M")
fn = Path.home() / "Downloads" / f"HoopsValue_Baseline_{stamp}.xlsx"
wb.save(fn)
print(f"OPEN THIS: {fn}")
print(f"  rosters: {len(roster_rows)} (kept options + guaranteed + picks) | free agents to assign: {len(grid)}")
